import openpyxl

def main():
    wb = openpyxl.load_workbook('sample.xlsx')
    ws = wb['Sheet1']

    for r in ws.rows:
        row_index = r[0].row
        if row_index == 1:
            ws.cell(row=row_index, column=5).value = 'Sum'
            ws.cell(row=row_index, column=6).value = 'Avg'
            continue

        math = r[1].value
        eng = r[2].value
        history = r[3].value
        summary = math + eng + history

        ws.cell(row=row_index, column=5).value = summary
        ws.cell(row=row_index, column=6).value = summary / 3
        ws.cell(row=row_index, column=6).number_format = '#.#'

    wb.save('sample2.xlsx')
    wb.close()

if __name__ == '__main__':
    main()