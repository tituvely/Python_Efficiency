import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color
from openpyxl.chart import BarChart, Reference, Series

def set_title(ws):
    ws.merge_cells('A1:E1')
    ws['A1'] = 'Report'

    title = ws['A1']
    title.font = Font(name='nanum gothic', size=18, bold=True)
    title.alignment = Alignment(horizontal='center', vertical='center')
    title.fill = PatternFill(patternType='solid', fgColor=Color('808080'))

    box = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )
    ws['A1'].border = box
    ws['B1'].border = box
    ws['C1'].border = box
    ws['D1'].border = box
    ws['E1'].border = box
   
def set_data(ws):
    rows = [
        ['V1', 'V2', 'V3', 'V4', 'V5'],
        [20, 90, 15, 80, 75],
        [30, 80, 25, 45, 80],
        [40, 70, 55, 88, 75],
        [50, 60, 75, 87, 89],
        [60, 50, 35, 85, 55],
        [70, 40, 5, 76, 85],
    ]

    for row in rows:
        ws.append(row)

def set_chart(ws):
    categories = Reference(ws, min_col=1, min_row=3, max_row=8)
    values = Reference(ws, min_col=1, min_row=2, max_col=5, max_row=8)

    chart = BarChart()
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(categories)
    ws.add_chart(chart, 'A10')

def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Report'

    set_title(ws)
    set_data(ws)
    set_chart(ws)

if __name__ == '__main__':
    main()